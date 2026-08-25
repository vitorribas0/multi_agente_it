"""Endpoints SSE que adaptam o Codex App Server ao contrato do frontend."""

from __future__ import annotations

import csv
import fcntl
import json
import queue
import re
import shutil
import threading
import time
import unicodedata
from dataclasses import dataclass, field
from datetime import timedelta
from pathlib import Path
from uuid import uuid4

from django.conf import settings
from django.core.exceptions import ValidationError
from django.db import IntegrityError, transaction
from django.http import HttpRequest, JsonResponse, StreamingHttpResponse
from django.utils import timezone
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_GET, require_http_methods

from .codex_app_server import CodexAppServer, CodexAppServerError
from .models import Conversation, Execution, ExecutionInteraction, Message, ToolCall
from .playbook_runtime import (
    normalize_execution_policy,
    ordered_playbook_nodes,
    playbook_plan,
    playbook_plan_explanation,
    playbook_stage_prompt,
    playbook_synthesis_prompt,
    set_plan_stage,
    snapshot_for_playbook,
)
from .views import _UNSET, _apply_playbook, _message_payload, _resolve_conversation
from tools.gerar_html import gerar_html


_HTML_TITLE_RE = re.compile(r"<title[^>]*>(.*?)</title>", re.IGNORECASE | re.DOTALL)
_HTML_EXPORT_PREFIX = "/api/exports/"
_CONVERSATION_ARTIFACT_PREFIX = "/api/conversations/"
_OUTPUT_DIRNAME = "saida"
_WORK_DIRNAME = "trabalho"
_EVIDENCE_DIRNAME = "evidencias"
_VERSIONS_DIRNAME = "versoes"
_LEGACY_ARTIFACTS_DIRNAME = "artefatos"
_GENERATED_ARTIFACT_EXTENSIONS = {".csv", ".xlsx", ".pdf", ".html", ".htm"}
_MAX_GENERATED_ARTIFACT_BYTES = 200 * 1024 * 1024
_MAX_GENERATED_ARTIFACTS_PER_TURN = 20
_TRACE_TEXT_LIMIT = 6000
_TRACE_RESULT_PREVIEW_LIMIT = 1200
_TRACE_META = {
    "commandExecution": ("codex_command", "⌨️"),
    "fileChange": ("codex_file_change", "📝"),
    "mcpToolCall": ("codex_mcp", "🔌"),
    "dynamicToolCall": ("codex_tool", "⚙️"),
    "collabToolCall": ("codex_subagent", "🤝"),
    "webSearch": ("codex_web_search", "🌐"),
    "imageView": ("codex_image_view", "🖼️"),
    "contextCompaction": ("codex_compaction", "🧠"),
}


def _normalize_live_plan(raw_plan: list[dict]) -> list[dict]:
    """Reduz o plano para o contrato público exibido no frontend."""
    return [
        {
            "step": _sanitize_trace_text(item.get("step"), 500),
            "status": item.get("status") or "pending",
        }
        for item in raw_plan[:20]
        if isinstance(item, dict)
    ]


def _advance_live_plan(plan: list[dict]) -> list[dict] | None:
    """Avança o plano quando uma ação concreta termina.

    O App Server pode não reenviar ``turn/plan/updated`` entre duas tools. Nesse
    caso, usar a conclusão da ação como marco evita uma interface congelada.
    Um novo plano explícito do App Server sempre substitui esta estimativa.
    """
    if not plan:
        return None
    next_plan = [dict(item) for item in plan]
    active_index = next((i for i, item in enumerate(next_plan) if item.get("status") == "inProgress"), None)
    if active_index is None:
        active_index = next((i for i, item in enumerate(next_plan) if item.get("status") == "pending"), None)
    if active_index is None:
        return None

    next_plan[active_index]["status"] = "completed"
    following = next((i for i, item in enumerate(next_plan) if item.get("status") == "pending"), None)
    if following is not None:
        next_plan[following]["status"] = "inProgress"
    return next_plan


@dataclass
class _PendingCodexInteraction:
    conversation_id: int
    method: str
    params: dict
    ready: threading.Event = field(default_factory=threading.Event)
    response: dict | None = None


_PENDING_INTERACTIONS: dict[str, _PendingCodexInteraction] = {}
_PENDING_INTERACTIONS_LOCK = threading.Lock()
_INTERACTION_TIMEOUT_SECONDS = 10 * 60
_CODEX_FORCE_STOP_SECONDS = 2.0
_EXECUTION_EVENT_LIMIT = 250
_WORKER_QUEUE_TIMEOUT_SECONDS = max(
    5, int(getattr(settings, "ATENA_WORKER_QUEUE_TIMEOUT_SECONDS", 20))
)
_CODEX_RUNTIME_ID = uuid4().hex


@dataclass
class _ActiveCodexExecution:
    conversation_id: int
    execution_id: str
    stop_event: threading.Event = field(default_factory=threading.Event)
    client: CodexAppServer | None = None
    force_timer: threading.Timer | None = None
    lock: threading.Lock = field(default_factory=threading.Lock)


_ACTIVE_CODEX_EXECUTIONS: dict[int, _ActiveCodexExecution] = {}
_ACTIVE_CODEX_EXECUTIONS_LOCK = threading.Lock()


class _CodexExecutionStopped(RuntimeError):
    pass


def _register_codex_execution(conversation_id: int, execution_id: str) -> _ActiveCodexExecution:
    execution = _ActiveCodexExecution(conversation_id, execution_id)
    with _ACTIVE_CODEX_EXECUTIONS_LOCK:
        previous = _ACTIVE_CODEX_EXECUTIONS.get(conversation_id)
        _ACTIVE_CODEX_EXECUTIONS[conversation_id] = execution
    if previous is not None:
        previous.stop_event.set()
        with previous.lock:
            previous_client = previous.client
        if previous_client is not None:
            try:
                previous_client.interrupt()
            except CodexAppServerError:
                previous_client.close()
    return execution


def _attach_codex_client(execution: _ActiveCodexExecution, client: CodexAppServer) -> None:
    with execution.lock:
        execution.client = client
        should_stop = execution.stop_event.is_set()
    if should_stop:
        client.interrupt()


def _release_pending_interactions(conversation_id: int) -> None:
    with _PENDING_INTERACTIONS_LOCK:
        pending_items = [
            pending
            for pending in _PENDING_INTERACTIONS.values()
            if pending.conversation_id == conversation_id and not pending.ready.is_set()
        ]
        for pending in pending_items:
            pending.response = _safe_interaction_fallback(pending.method)
            pending.ready.set()
    now = timezone.now()
    durable_items = ExecutionInteraction.objects.filter(
        execution__conversation_id=conversation_id,
        status="pending",
    )
    for durable in durable_items:
        ExecutionInteraction.objects.filter(
            pk=durable.pk,
            status="pending",
        ).update(
            status="cancelled",
            response=_safe_interaction_fallback(durable.method),
            responded_at=now,
            updated_at=now,
        )


def _force_stop_codex_execution(execution: _ActiveCodexExecution) -> None:
    with _ACTIVE_CODEX_EXECUTIONS_LOCK:
        if _ACTIVE_CODEX_EXECUTIONS.get(execution.conversation_id) is not execution:
            return
    with execution.lock:
        client = execution.client
    if client is not None:
        client.close()


def request_codex_stop(conversation_id: int, execution_id: str | None = None) -> bool:
    """Interrompe o turno Codex ativo e arma um encerramento de segurança."""
    with _ACTIVE_CODEX_EXECUTIONS_LOCK:
        execution = _ACTIVE_CODEX_EXECUTIONS.get(conversation_id)
    if execution is None or (execution_id and execution.execution_id != execution_id):
        return False

    execution.stop_event.set()
    _release_pending_interactions(conversation_id)
    with execution.lock:
        client = execution.client
        timer = execution.force_timer
    if client is not None:
        try:
            client.interrupt()
        except CodexAppServerError:
            client.close()
    if timer is None:
        force_timer = threading.Timer(
            _CODEX_FORCE_STOP_SECONDS,
            _force_stop_codex_execution,
            args=(execution,),
        )
        force_timer.daemon = True
        with execution.lock:
            if execution.force_timer is None:
                execution.force_timer = force_timer
                force_timer.start()
    return True


def _execution_public_payload(execution: Execution) -> dict:
    return {
        "id": str(execution.id),
        "conversation_id": execution.conversation_id,
        "engine": execution.engine,
        "backend": execution.backend,
        "status": execution.status,
        "events": execution.events or [],
        "plan": execution.plan or [],
        "plan_explanation": execution.plan_explanation,
        "error": execution.error,
        "stop_requested_at": (
            execution.stop_requested_at.isoformat() if execution.stop_requested_at else None
        ),
        "started_at": execution.started_at.isoformat() if execution.started_at else None,
        "finished_at": execution.finished_at.isoformat() if execution.finished_at else None,
        "last_heartbeat_at": (
            execution.last_heartbeat_at.isoformat() if execution.last_heartbeat_at else None
        ),
        "created_at": execution.created_at.isoformat(),
        "updated_at": execution.updated_at.isoformat(),
    }


def _local_worker_is_running() -> bool:
    """Confirma pelo lock exclusivo se o worker local continua vivo."""
    lock_path = Path(settings.BASE_DIR) / "runtime" / "agent-worker.lock"
    if not lock_path.parent.is_dir():
        return False
    try:
        with lock_path.open("a+") as lock_file:
            try:
                fcntl.flock(lock_file.fileno(), fcntl.LOCK_EX | fcntl.LOCK_NB)
            except BlockingIOError:
                return True
            fcntl.flock(lock_file.fileno(), fcntl.LOCK_UN)
            return False
    except OSError:
        # Em caso de dúvida não encerramos uma execução potencialmente válida.
        return True


def _expire_unclaimed_worker_execution(execution: Execution) -> bool:
    """Encerra uma fila que nenhum worker assumiu dentro do limite local."""
    if execution.backend != "local-worker" or execution.status != "queued":
        return False
    deadline = timezone.now() - timedelta(seconds=_WORKER_QUEUE_TIMEOUT_SECONDS)
    if execution.created_at > deadline or _local_worker_is_running():
        return False
    message = (
        "O worker da Atena não iniciou a execução. Inicie "
        "'python manage.py run_agent_worker' e tente novamente."
    )
    changed = Execution.objects.filter(
        pk=execution.pk,
        backend="local-worker",
        status="queued",
        created_at__lte=deadline,
    ).update(
        status="failed",
        error=message,
        finished_at=timezone.now(),
        last_heartbeat_at=timezone.now(),
        updated_at=timezone.now(),
    )
    return bool(changed)


def _recover_orphaned_local_executions(conversation_id: int | None = None) -> int:
    """Fecha execuções locais que pertencem a outro processo Django.

    Em produção, backends ECS/SQS terão sua própria estratégia de reconciliação;
    esta regra é deliberadamente restrita ao backend local.
    """
    query = Execution.objects.filter(
        backend="local",
        status__in=Execution.ACTIVE_STATUSES,
    ).exclude(runtime_id=_CODEX_RUNTIME_ID)
    if conversation_id is not None:
        query = query.filter(conversation_id=conversation_id)
    now = timezone.now()
    return query.update(
        status="failed",
        error="Execução local encerrada porque o processo do servidor foi reiniciado.",
        finished_at=now,
        last_heartbeat_at=now,
    )


def _persist_execution_event(execution_id, event: dict) -> int | None:
    """Persiste somente o contrato público necessário para reconectar a UI."""
    try:
        execution = Execution.objects.get(pk=execution_id)
        event_type = event.get("type")
        stored_event: dict | None = None
        now = timezone.now()
        updates: dict = {"last_heartbeat_at": now, "updated_at": now}
        live_status: str | None = None
        if event_type in {"progress", "plan", "interaction", "interaction_resolved"}:
            stored_event = dict(event)
        elif event_type == "done":
            payload = event.get("payload") or {}
            stopped = bool(payload.get("stopped"))
            stored_event = {"type": "done", "stopped": stopped}
            updates.update({
                "status": "stopped" if stopped else "completed",
                "finished_at": timezone.now(),
                "error": "",
            })
        elif event_type == "error":
            message = _sanitize_trace_text(event.get("message"), 4000)
            stored_event = {"type": "error", "message": message}
            updates.update({
                "status": "failed",
                "finished_at": timezone.now(),
                "error": message,
            })

        if event_type == "plan":
            updates["plan"] = event.get("plan") or []
            updates["plan_explanation"] = _sanitize_trace_text(
                event.get("explanation"), 1200
            )
        elif event_type == "interaction":
            live_status = "waiting_user"
        elif event_type == "progress":
            live_status = "running"

        if stored_event is not None:
            previous_events = execution.events or []
            try:
                previous_event = previous_events[-1] if previous_events else {}
                previous_sequence = int((previous_event or {}).get("sequence", 0))
            except (AttributeError, TypeError, ValueError):
                previous_sequence = 0
            stored_event["sequence"] = previous_sequence + 1
            updates["events"] = [*(execution.events or []), stored_event][-_EXECUTION_EVENT_LIMIT:]
        Execution.objects.filter(pk=execution_id).update(**updates)
        if live_status is not None:
            # O pedido de parada pode chegar entre a leitura e a gravação do
            # evento. A atualização condicional garante que progresso tardio
            # nunca reative uma execução que já está sendo interrompida.
            Execution.objects.filter(
                pk=execution_id,
                status__in=("queued", "starting", "running", "waiting_user"),
            ).update(status=live_status, updated_at=now)
        return stored_event.get("sequence") if stored_event is not None else None
    except Exception:
        # Falha de telemetria não pode derrubar o turno do agente.
        return None


class _ExecutionEventQueue(queue.Queue):
    def __init__(self, execution_id):
        super().__init__()
        self.execution_id = execution_id

    def put(self, item, block=True, timeout=None):
        if isinstance(item, dict) and item.get("type"):
            sequence = _persist_execution_event(self.execution_id, item)
            if sequence is not None:
                item = {**item, "sequence": sequence}
        return super().put(item, block=block, timeout=timeout)


def request_codex_execution_stop(execution: Execution) -> bool:
    if execution.status in Execution.TERMINAL_STATUSES:
        return False
    now = timezone.now()

    # O backend durável roda em outro processo. O pedido fica no banco e é
    # observado pelo worker; tentar acessar o registry em memória da API faria
    # uma execução válida parecer órfã.
    if execution.backend == "local-worker":
        if execution.status == "queued":
            changed = Execution.objects.filter(
                pk=execution.id,
                status="queued",
            ).update(
                status="stopped",
                stop_requested_at=now,
                finished_at=now,
                last_heartbeat_at=now,
                updated_at=now,
            )
            if not changed:
                return False
            _release_pending_interactions(execution.conversation_id)
            _persist_execution_event(execution.id, {
                "type": "done",
                "payload": {"stopped": True},
            })
            return True

        changed = Execution.objects.filter(
            pk=execution.id,
            status__in=("starting", "running", "waiting_user"),
        ).update(
            status="stopping",
            stop_requested_at=now,
            last_heartbeat_at=now,
            updated_at=now,
        )
        if changed:
            _release_pending_interactions(execution.conversation_id)
        return bool(changed)

    Execution.objects.filter(pk=execution.id).update(
        status="stopping",
        stop_requested_at=now,
        last_heartbeat_at=now,
    )
    stopped = request_codex_stop(execution.conversation_id, str(execution.id))
    if not stopped:
        Execution.objects.filter(pk=execution.id, status="stopping").update(
            status="failed",
            error="A execução não estava mais vinculada ao processo local.",
            finished_at=timezone.now(),
        )
    return stopped


def _watch_persisted_stop(
    execution_record: Execution,
    active_execution: _ActiveCodexExecution,
    finished: threading.Event,
) -> None:
    """Traduz o estado de parada do banco para o cliente Codex do worker."""
    from django.db import close_old_connections

    close_old_connections()
    try:
        while not finished.wait(0.25):
            status = Execution.objects.filter(pk=execution_record.pk).values_list(
                "status", flat=True
            ).first()
            if status in {"stopping", "stopped", "failed"}:
                request_codex_stop(
                    execution_record.conversation_id,
                    str(execution_record.id),
                )
                return
            if status is None or status in Execution.TERMINAL_STATUSES:
                return
    finally:
        close_old_connections()


def _unregister_codex_execution(execution: _ActiveCodexExecution) -> None:
    with _ACTIVE_CODEX_EXECUTIONS_LOCK:
        if _ACTIVE_CODEX_EXECUTIONS.get(execution.conversation_id) is execution:
            _ACTIVE_CODEX_EXECUTIONS.pop(execution.conversation_id, None)
    with execution.lock:
        timer = execution.force_timer
        execution.force_timer = None
        execution.client = None
    if timer is not None:
        timer.cancel()


def _interaction_public_payload(token: str, method: str, params: dict) -> dict:
    """Converte uma request do App Server num contrato pequeno para a UI."""
    base = {
        "token": token,
        "reason": _sanitize_trace_text(params.get("reason"), 1000),
    }
    if method in {"item/tool/requestUserInput", "tool/requestUserInput"}:
        questions = []
        for raw in (params.get("questions") or [])[:3]:
            if not isinstance(raw, dict):
                continue
            questions.append({
                "id": str(raw.get("id") or "")[:120],
                "header": _sanitize_trace_text(raw.get("header"), 80),
                "question": _sanitize_trace_text(raw.get("question"), 1000),
                "isOther": bool(raw.get("isOther")),
                "isSecret": bool(raw.get("isSecret")),
                "options": [
                    {
                        "label": _sanitize_trace_text(option.get("label"), 120),
                        "description": _sanitize_trace_text(option.get("description"), 400),
                    }
                    for option in (raw.get("options") or [])[:8]
                    if isinstance(option, dict)
                ],
            })
        return {
            **base,
            "kind": "question",
            "title": "O agente precisa de você",
            "questions": questions,
        }
    if method == "item/commandExecution/requestApproval":
        network = params.get("networkApprovalContext") or {}
        available = [
            value for value in (params.get("availableDecisions") or [])
            if isinstance(value, str)
        ]
        return {
            **base,
            "kind": "command_approval",
            "title": "Autorizar comando?",
            "command": _sanitize_trace_text(params.get("command"), 4000),
            "cwd": Path(str(params.get("cwd") or ".")).name,
            "network": {
                "host": _sanitize_trace_text(network.get("host"), 300),
                "protocol": _sanitize_trace_text(network.get("protocol"), 40),
            } if isinstance(network, dict) and network else None,
            "availableDecisions": available,
        }
    if method == "item/fileChange/requestApproval":
        return {
            **base,
            "kind": "file_approval",
            "title": "Autorizar alteração de arquivo?",
            "grantRoot": _sanitize_trace_text(params.get("grantRoot"), 1000),
        }
    return {
        **base,
        "kind": "permission_approval",
        "title": "Autorizar permissões extras?",
        "cwd": Path(str(params.get("cwd") or ".")).name,
        "permissions": _bounded_trace_value(params.get("permissions") or {}),
    }


def _safe_interaction_fallback(method: str) -> dict:
    if method == "item/permissions/requestApproval":
        return {"permissions": {}, "scope": "turn"}
    if method in {"item/tool/requestUserInput", "tool/requestUserInput"}:
        return {"answers": {}}
    return {"decision": "decline"}


def _approve_interaction_for_turn(method: str, params: dict) -> dict:
    """Aprova uma solicitação sem ampliar o escopo além do turno atual."""
    if method == "item/permissions/requestApproval":
        return {
            "permissions": params.get("permissions") or {},
            "scope": "turn",
        }
    if method in {"item/tool/requestUserInput", "tool/requestUserInput"}:
        return {"answers": {}}
    return {"decision": "accept"}


def _interaction_response_from_data(method: str, params: dict, data: dict) -> tuple[dict | None, str | None, int]:
    if data.get("approve_all") is True:
        if method in {"item/tool/requestUserInput", "tool/requestUserInput"}:
            return None, "Perguntas precisam ser respondidas.", 400
        return {
            **_approve_interaction_for_turn(method, params),
            "__approve_all_for_turn__": True,
        }, None, 200
    if data.get("cancel") is True:
        return _safe_interaction_fallback(method), None, 200
    if method in {"item/tool/requestUserInput", "tool/requestUserInput"}:
        raw_answers = data.get("answers") or {}
        if not isinstance(raw_answers, dict):
            return None, "Respostas inválidas.", 400
        valid_ids = {
            str(question.get("id"))
            for question in (params.get("questions") or [])
            if isinstance(question, dict) and question.get("id") is not None
        }
        answers = {}
        for question_id in valid_ids:
            values = raw_answers.get(question_id, [])
            if isinstance(values, str):
                values = [values]
            if not isinstance(values, list):
                values = []
            cleaned = [
                str(value).strip()[:4000]
                for value in values[:10]
                if str(value).strip()
            ]
            if cleaned:
                answers[question_id] = {"answers": cleaned}
        if valid_ids and len(answers) != len(valid_ids):
            return None, "Responda todas as perguntas.", 400
        return {"answers": answers}, None, 200
    if method == "item/permissions/requestApproval":
        approved = bool(data.get("approve"))
        scope = "session" if data.get("scope") == "session" else "turn"
        return {
            "permissions": (params.get("permissions") or {}) if approved else {},
            "scope": scope,
        }, None, 200
    decision = data.get("decision")
    if decision not in {"accept", "acceptForSession", "decline", "cancel"}:
        return None, "Decisão inválida.", 400
    return {"decision": decision}, None, 200


def _wait_for_durable_interaction(
    conv,
    execution_record: Execution,
    events: "queue.Queue[dict]",
    method: str,
    params: dict,
) -> dict:
    timeout = _INTERACTION_TIMEOUT_SECONDS
    auto_resolution_ms = params.get("autoResolutionMs")
    if isinstance(auto_resolution_ms, (int, float)) and auto_resolution_ms > 0:
        timeout = min(timeout, max(1, auto_resolution_ms / 1000))
    now = timezone.now()
    interaction = ExecutionInteraction.objects.create(
        execution=execution_record,
        method=method,
        params=params,
        expires_at=now + timedelta(seconds=timeout),
    )
    token = str(interaction.token)
    Conversation.objects.filter(pk=conv.id).update(awaiting_human_input=True)
    events.put({
        "type": "interaction",
        "interaction": _interaction_public_payload(token, method, params),
    })

    deadline = time.monotonic() + timeout
    response = None
    resolution = "expired"
    while time.monotonic() < deadline:
        snapshot = ExecutionInteraction.objects.filter(pk=interaction.pk).values(
            "status", "response"
        ).first()
        if not snapshot:
            break
        if snapshot["status"] in {"responded", "cancelled"}:
            response = snapshot["response"] or _safe_interaction_fallback(method)
            resolution = snapshot["status"]
            break
        execution_status = Execution.objects.filter(pk=execution_record.pk).values_list(
            "status", flat=True
        ).first()
        if execution_status in {"stopping", "stopped", "failed"}:
            response = _safe_interaction_fallback(method)
            resolution = "cancelled"
            ExecutionInteraction.objects.filter(
                pk=interaction.pk,
                status="pending",
            ).update(
                status="cancelled",
                response=response,
                responded_at=timezone.now(),
                updated_at=timezone.now(),
            )
            break
        time.sleep(0.2)

    if response is None:
        response = _safe_interaction_fallback(method)
        ExecutionInteraction.objects.filter(
            pk=interaction.pk,
            status="pending",
        ).update(
            status="expired",
            response=response,
            responded_at=timezone.now(),
            updated_at=timezone.now(),
        )
    Conversation.objects.filter(pk=conv.id).update(awaiting_human_input=False)
    events.put({"type": "interaction_resolved", "token": token})
    if resolution == "expired":
        events.put({
            "type": "progress",
            "stage": "thinking",
            "icon": "⏱️",
            "text": "Solicitação expirada; a ação não foi autorizada",
        })
    elif resolution == "responded":
        events.put({
            "type": "progress",
            "stage": "thinking",
            "icon": "▶️",
            "text": "Resposta recebida; retomando a execução",
        })
    return response


def _wait_for_interaction(
    conv,
    events: "queue.Queue[dict]",
    method: str,
    params: dict,
    execution_record: Execution | None = None,
) -> dict:
    if execution_record is not None:
        return _wait_for_durable_interaction(
            conv,
            execution_record,
            events,
            method,
            params,
        )
    token = uuid4().hex
    pending = _PendingCodexInteraction(conv.id, method, params)
    with _PENDING_INTERACTIONS_LOCK:
        _PENDING_INTERACTIONS[token] = pending

    Conversation.objects.filter(pk=conv.id).update(awaiting_human_input=True)
    events.put({
        "type": "interaction",
        "interaction": _interaction_public_payload(token, method, params),
    })
    timeout = _INTERACTION_TIMEOUT_SECONDS
    auto_resolution_ms = params.get("autoResolutionMs")
    if isinstance(auto_resolution_ms, (int, float)) and auto_resolution_ms > 0:
        timeout = min(timeout, max(1, auto_resolution_ms / 1000))
    resolved = pending.ready.wait(timeout=timeout)

    with _PENDING_INTERACTIONS_LOCK:
        _PENDING_INTERACTIONS.pop(token, None)
    Conversation.objects.filter(pk=conv.id).update(awaiting_human_input=False)
    events.put({"type": "interaction_resolved", "token": token})
    if not resolved or pending.response is None:
        events.put({
            "type": "progress",
            "stage": "thinking",
            "icon": "⏱️",
            "text": "Solicitação expirada; a ação não foi autorizada",
        })
        return _safe_interaction_fallback(method)
    events.put({
        "type": "progress",
        "stage": "thinking",
        "icon": "▶️",
        "text": "Resposta recebida; retomando a execução",
    })
    return pending.response


@csrf_exempt
def codex_interaction_respond(request, token: str):
    if request.method != "POST":
        return JsonResponse({"status": "error", "message": "Method not allowed"}, status=405)
    try:
        data = json.loads(request.body or "{}")
    except json.JSONDecodeError:
        return JsonResponse({"status": "error", "message": "JSON inválido."}, status=400)

    durable = None
    try:
        durable = ExecutionInteraction.objects.filter(pk=token).first()
    except (ValidationError, ValueError):
        durable = None
    if durable is not None:
        if durable.status != "pending":
            return JsonResponse({"status": "error", "message": "Solicitação já encerrada."}, status=409)
        response, error, status_code = _interaction_response_from_data(
            durable.method,
            durable.params or {},
            data,
        )
        if error:
            return JsonResponse({"status": "error", "message": error}, status=status_code)
        now = timezone.now()
        changed = ExecutionInteraction.objects.filter(
            pk=durable.pk,
            status="pending",
        ).update(
            status="responded",
            response=response or {},
            responded_at=now,
            updated_at=now,
        )
        if not changed:
            return JsonResponse({"status": "error", "message": "Solicitação já respondida."}, status=409)
        return JsonResponse({"status": "success"})

    with _PENDING_INTERACTIONS_LOCK:
        pending = _PENDING_INTERACTIONS.get(token)
        if pending is None:
            return JsonResponse(
                {"status": "error", "message": "Solicitação expirada ou inexistente."},
                status=404,
            )
        if pending.response is not None:
            return JsonResponse({"status": "error", "message": "Solicitação já respondida."}, status=409)

        response, error, status_code = _interaction_response_from_data(
            pending.method,
            pending.params,
            data,
        )
        if error:
            return JsonResponse({"status": "error", "message": error}, status=status_code)
        pending.response = response
        pending.ready.set()

    return JsonResponse({"status": "success"})


def _sanitize_trace_text(value, limit: int = _TRACE_TEXT_LIMIT) -> str:
    if value is None:
        return ""
    if not isinstance(value, str):
        try:
            value = json.dumps(value, ensure_ascii=False, default=str)
        except (TypeError, ValueError):
            value = str(value)
    value = re.sub(
        r"(?i)((?:OPENAI|AWS|AZURE|IARA)[A-Z0-9_]*(?:KEY|TOKEN|SECRET|PASSWORD)\s*[=:]\s*)([^\s,;]+)",
        r"\1[REDACTED]",
        value,
    )
    value = re.sub(r"\bsk-[A-Za-z0-9_-]{16,}\b", "[REDACTED]", value)
    return value if len(value) <= limit else value[:limit] + "…"


def _bounded_trace_value(value):
    """Mantém args estruturados pequenos e transforma payloads grandes em prévia."""
    if value is None or isinstance(value, (bool, int, float)):
        return value
    if isinstance(value, str):
        return _sanitize_trace_text(value, 4000)
    serialized = _sanitize_trace_text(value, 8000)
    try:
        parsed = json.loads(serialized)
    except json.JSONDecodeError:
        return serialized
    return parsed


def _codex_trace_args(item: dict) -> dict:
    item_type = item.get("type")
    if item_type == "commandExecution":
        return {
            "command": _sanitize_trace_text(item.get("command"), 4000),
            "cwd": Path(str(item.get("cwd") or ".")).name,
        }
    if item_type == "mcpToolCall":
        return {
            "server": item.get("server") or "",
            "tool": item.get("tool") or "",
            "arguments": _bounded_trace_value(item.get("arguments") or {}),
        }
    if item_type == "dynamicToolCall":
        return {
            "tool": item.get("tool") or "",
            "arguments": _bounded_trace_value(item.get("arguments") or {}),
        }
    if item_type == "collabToolCall":
        return {
            "tool": item.get("tool") or "",
            "prompt": _sanitize_trace_text(item.get("prompt"), 3000),
        }
    if item_type == "webSearch":
        return {
            "query": _sanitize_trace_text(item.get("query"), 2000),
            "action": _bounded_trace_value(item.get("action") or {}),
        }
    if item_type == "fileChange":
        return {
            "changes": [
                {"path": change.get("path"), "kind": change.get("kind")}
                for change in (item.get("changes") or [])[:50]
                if isinstance(change, dict)
            ]
        }
    if item_type == "imageView":
        return {"path": Path(str(item.get("path") or "imagem")).name}
    return {}


def _codex_trace_label(item: dict) -> str:
    item_type = item.get("type")
    if item_type == "commandExecution":
        command = _sanitize_trace_text(item.get("command"), 100).splitlines()[0]
        return f"Executando no sandbox · {command}" if command else "Executando no sandbox"
    if item_type == "mcpToolCall":
        target = " · ".join(filter(None, [str(item.get("server") or ""), str(item.get("tool") or "")]))
        return f"Consultando ferramenta · {target}" if target else "Consultando ferramenta"
    if item_type == "dynamicToolCall":
        return f"Executando ferramenta · {item.get('tool') or 'Atena'}"
    if item_type == "collabToolCall":
        return f"Coordenando subagente · {item.get('tool') or 'Atena'}"
    if item_type == "webSearch":
        query = _sanitize_trace_text(item.get("query"), 100)
        return f"Pesquisando na web · {query}" if query else "Pesquisando na web"
    if item_type == "fileChange":
        return f"Preparando alterações em {len(item.get('changes') or [])} arquivo(s)"
    if item_type == "imageView":
        return f"Inspecionando imagem · {Path(str(item.get('path') or 'imagem')).name}"
    if item_type == "contextCompaction":
        return "Organizando o contexto da conversa"
    return "Executando atividade da Atena"


def _codex_trace_result(item: dict, streamed_output: str = "") -> tuple[str, str]:
    item_type = item.get("type")
    status = str(item.get("status") or "completed").lower()
    error_value = item.get("error")
    error = _sanitize_trace_text(error_value)
    if status in {"failed", "declined", "cancelled", "canceled"} and not error:
        error = f"Atividade encerrada com status {status}."

    if item_type == "commandExecution":
        result = item.get("aggregatedOutput") or streamed_output
        exit_code = item.get("exitCode")
        if exit_code not in (None, 0) and not error:
            error = f"Comando encerrado com código {exit_code}."
        return _sanitize_trace_text(result), error
    if item_type == "mcpToolCall":
        return _sanitize_trace_text(item.get("result")), error
    if item_type == "dynamicToolCall":
        return _sanitize_trace_text(item.get("contentItems") or item.get("success")), error
    if item_type == "collabToolCall":
        return _sanitize_trace_text(item.get("agentStatus") or status), error
    if item_type == "webSearch":
        return "Pesquisa concluída.", error
    if item_type == "fileChange":
        return f"{len(item.get('changes') or [])} alteração(ões) processada(s).", error
    if item_type == "imageView":
        return "Imagem inspecionada.", error
    if item_type == "contextCompaction":
        return "Contexto da conversa organizado.", error
    return _sanitize_trace_text(status), error


def _codex_trace_record(item: dict, streamed_output: str = "", elapsed_ms: int = 0) -> dict | None:
    item_type = item.get("type")
    meta = _TRACE_META.get(item_type)
    item_id = str(item.get("id") or "")
    if not meta or not item_id:
        return None
    result, error = _codex_trace_result(item, streamed_output)
    duration = item.get("durationMs")
    try:
        duration_ms = int(duration if duration is not None else elapsed_ms)
    except (TypeError, ValueError):
        duration_ms = elapsed_ms
    return {
        "id": item_id,
        "item_type": item_type,
        "tool": meta[0],
        "icon": meta[1],
        "label": _codex_trace_label(item),
        "args": _codex_trace_args(item),
        "result": result,
        "error": error,
        "duration_ms": max(0, duration_ms),
    }


def _artifact_snapshot(artifacts_dir: Path) -> dict[str, tuple[int, int]]:
    """Fotografa somente formatos publicáveis antes de iniciar o turno."""
    artifacts_dir.mkdir(parents=True, exist_ok=True)
    snapshot: dict[str, tuple[int, int]] = {}
    for path in artifacts_dir.rglob("*"):
        if path.is_symlink() or not path.is_file():
            continue
        if path.suffix.lower() not in _GENERATED_ARTIFACT_EXTENSIONS:
            continue
        try:
            stat = path.stat()
            relative = str(path.relative_to(artifacts_dir))
        except (OSError, ValueError):
            continue
        snapshot[relative] = (stat.st_mtime_ns, stat.st_size)
    return snapshot


def _csv_dimensions(path: Path) -> tuple[int | None, int | None]:
    try:
        with path.open("r", encoding="utf-8-sig", errors="replace", newline="") as handle:
            sample = handle.read(8192)
            handle.seek(0)
            try:
                dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
            except csv.Error:
                dialect = csv.excel
            reader = csv.reader(handle, dialect)
            header = next(reader, [])
            rows = sum(1 for _ in reader)
        return rows, len(header)
    except (OSError, UnicodeError, csv.Error):
        return None, None


def _xlsx_dimensions(path: Path) -> tuple[int | None, int | None, list[str]]:
    try:
        from openpyxl import load_workbook
        from openpyxl.utils.cell import range_boundaries

        workbook = load_workbook(path, read_only=True, data_only=True)
        sheets = list(workbook.sheetnames)
        dimensions: list[tuple[int, int]] = []
        for sheet in workbook.worksheets:
            max_row = sheet.max_row
            max_column = sheet.max_column
            if max_row is None or max_column is None:
                try:
                    _min_col, _min_row, max_column, max_row = range_boundaries(
                        sheet.calculate_dimension(force=True)
                    )
                except (TypeError, ValueError):
                    max_row, max_column = 0, 0
            dimensions.append((max_row or 0, max_column or 0))
        rows = sum(max(0, max_row - 1) for max_row, _max_column in dimensions)
        columns = max((max_column for _max_row, max_column in dimensions), default=0)
        workbook.close()
        return rows, columns, sheets[:30]
    except Exception:
        return None, None, []


def _pdf_pages(path: Path) -> int | None:
    try:
        from pypdf import PdfReader

        return len(PdfReader(str(path)).pages)
    except Exception:
        return None


def _conversation_artifacts_dir(conversation_id: int) -> Path:
    target = Path(settings.BASE_DIR) / "runtime" / "codex_sessions" / str(conversation_id) / _OUTPUT_DIRNAME
    target.mkdir(parents=True, exist_ok=True)
    return target


def _conversation_artifact_url(conversation_id: int, filename: str) -> str:
    return f"/api/conversations/{conversation_id}/artifacts/{filename}"


def _generated_artifact_attachment(path: Path, conversation_id: int) -> dict | None:
    """Publica a saída final e arquiva a versão anterior, quando existente."""
    try:
        size = path.stat().st_size
    except OSError:
        return None
    if size <= 0 or size > _MAX_GENERATED_ARTIFACT_BYTES:
        return None

    source_suffix = path.suffix.lower()
    if source_suffix not in _GENERATED_ARTIFACT_EXTENSIONS:
        return None
    formato = "html" if source_suffix == ".htm" else source_suffix.lstrip(".")
    safe_stem = re.sub(r"[^A-Za-z0-9_-]+", "_", path.stem).strip("_")[:50]
    safe_stem = safe_stem or "artefato"
    published_filename = f"{safe_stem}.{formato}"
    output_path = _conversation_artifacts_dir(conversation_id) / published_filename
    try:
        if path.resolve() != output_path.resolve():
            if output_path.exists():
                versions_dir = (
                    output_path.parent.parent / _VERSIONS_DIRNAME / safe_stem
                )
                versions_dir.mkdir(parents=True, exist_ok=True)
                archived = versions_dir / f"{safe_stem}_{time.time_ns()}_{uuid4().hex[:6]}.{formato}"
                shutil.copy2(output_path, archived)
            shutil.copy2(path, output_path)
    except OSError:
        return None

    payload: dict = {
        "kind": "export",
        "ok": True,
        "filename": published_filename,
        "download_url": _conversation_artifact_url(conversation_id, published_filename),
        "formato": formato,
        "titulo": path.stem,
        "size_kb": round(size / 1024, 1),
    }
    if formato == "csv":
        payload["linhas"], payload["colunas"] = _csv_dimensions(path)
    elif formato == "xlsx":
        rows, columns, sheets = _xlsx_dimensions(path)
        payload.update({"linhas": rows, "colunas": columns, "abas": sheets})
    elif formato == "pdf":
        payload["paginas"] = _pdf_pages(path)
    return payload


def _collect_generated_artifacts(
    artifacts_dir: Path,
    before: dict[str, tuple[int, int]],
    conversation_id: int,
) -> list[dict]:
    """Publica arquivos novos ou alterados pelo Codex durante o turno."""
    root = artifacts_dir.resolve()
    changed: list[tuple[int, str, Path]] = []
    for path in artifacts_dir.rglob("*"):
        if path.is_symlink() or not path.is_file():
            continue
        if path.suffix.lower() not in _GENERATED_ARTIFACT_EXTENSIONS:
            continue
        try:
            resolved = path.resolve(strict=True)
            resolved.relative_to(root)
            stat = resolved.stat()
            relative = str(resolved.relative_to(root))
        except (OSError, ValueError):
            continue
        signature = (stat.st_mtime_ns, stat.st_size)
        if before.get(relative) != signature:
            changed.append((stat.st_mtime_ns, relative, resolved))

    attachments: list[dict] = []
    for _mtime, _relative, path in sorted(changed)[:_MAX_GENERATED_ARTIFACTS_PER_TURN]:
        attachment = _generated_artifact_attachment(path, conversation_id)
        if attachment:
            attachments.append(attachment)
    return attachments


def _generated_artifacts_manifest(workspace: Path) -> list[dict]:
    artifacts_dir = workspace / _OUTPUT_DIRNAME
    items: list[dict] = []
    for relative, (_mtime, size) in sorted(_artifact_snapshot(artifacts_dir).items()):
        items.append({
            "arquivo": str(Path(_OUTPUT_DIRNAME) / relative),
            "formato": Path(relative).suffix.lower().lstrip("."),
            "tamanho_bytes": size,
        })
    return items


def _directory_file_manifest(root: Path) -> list[str]:
    """Lista arquivos regulares de uma área interna, sem seguir symlinks."""
    if not root.is_dir():
        return []
    files: list[str] = []
    for path in root.rglob("*"):
        if path.is_file() and not path.is_symlink():
            try:
                files.append(str(path.relative_to(root)))
            except ValueError:
                continue
    return sorted(files)


def _strip_codex_file_citations(answer: str) -> str:
    """O frontend já renderiza o card; remove a diretiva interna do Codex."""
    cleaned = re.sub(r":codex-file-citation\{[^{}]*\}", "", answer)
    return re.sub(r"[ \t]+(?=\n|$)", "", cleaned).strip()


def _split_html_response(answer: str) -> tuple[str, str | None]:
    """Separa texto explicativo de um documento HTML completo retornado inline."""
    lowered = answer.lower()
    starts = [pos for pos in (lowered.find("<!doctype html"), lowered.find("<html")) if pos >= 0]
    if not starts:
        return answer, None
    start = min(starts)
    end = lowered.rfind("</html>")
    if end < start:
        return answer, None
    end += len("</html>")

    html = answer[start:end].strip()
    before = answer[:start].strip()
    after = answer[end:].strip()
    # Remove somente as cercas externas que normalmente envolvem o documento.
    before = re.sub(r"```(?:html|htm)?\s*$", "", before, flags=re.IGNORECASE).strip()
    after = re.sub(r"^```", "", after).strip()
    visible_text = "\n\n".join(part for part in (before, after) if part).strip()
    return visible_text, html


def _materialize_html_response(answer: str, conversation_id: int) -> tuple[str, list[dict]]:
    """Converte HTML inline do Codex no attachment já entendido pelo Angular."""
    visible_text, html = _split_html_response(answer)
    if html is None:
        return answer, []

    title_match = _HTML_TITLE_RE.search(html)
    title = re.sub(r"\s+", " ", title_match.group(1)).strip() if title_match else "Relatório"
    session: dict = {"__conversation_id": conversation_id}
    raw_result = gerar_html(
        _session=session,
        html=html,
        titulo=title or "Relatório",
        nome_arquivo="relatorio_codex",
    )
    try:
        result = json.loads(raw_result)
    except (TypeError, json.JSONDecodeError):
        result = {"erro": "Resposta inválida ao salvar HTML."}
    if result.get("erro"):
        return answer, []

    attachments = session.get("__pending_attachments") or []
    if not visible_text:
        visible_text = "Relatório HTML gerado. Use **Visualizar** para abrir o painel interativo."
    return visible_text, attachments


def _normalize_request(text: str) -> str:
    normalized = unicodedata.normalize("NFKD", text.lower())
    return "".join(char for char in normalized if not unicodedata.combining(char))


def _is_html_revision_request(text: str) -> bool:
    """Reconhece referências explícitas a um artefato HTML já criado."""
    normalized = _normalize_request(text)
    artifact = r"(?:html|relatorio|dashboard|pagina|artefato)"
    reference = r"(?:nesse|neste|naquele|no\s+mesmo|o\s+mesmo|mesmo|anterior|acima|existente|ultimo)"
    return bool(
        re.search(rf"\b{reference}\b.{{0,60}}\b{artifact}\b", normalized)
        or re.search(rf"\b{artifact}\b.{{0,60}}\b{reference}\b", normalized)
    )


def _latest_html_artifact(conv) -> str | None:
    """Lê com segurança o HTML do attachment mais recente da conversa."""
    export_dir = (Path(settings.BASE_DIR) / "exports").resolve()
    output_dir = _conversation_artifacts_dir(conv.id).resolve()
    legacy_dir = (output_dir.parent / _LEGACY_ARTIFACTS_DIRNAME).resolve()
    for message in conv.messages.order_by("-created_at"):
        for attachment in reversed(message.attachments or []):
            if attachment.get("kind") != "export" or attachment.get("formato") != "html":
                continue
            download_url = str(attachment.get("download_url") or "")
            candidate: Path | None = None
            local_prefix = f"{_CONVERSATION_ARTIFACT_PREFIX}{conv.id}/artifacts/"
            if download_url.startswith(local_prefix):
                filename = Path(download_url.removeprefix(local_prefix)).name
                candidate = (output_dir / filename).resolve()
                if candidate.parent != output_dir or not candidate.is_file():
                    candidate = (legacy_dir / filename).resolve()
                    if candidate.parent != legacy_dir:
                        continue
            elif download_url.startswith(_HTML_EXPORT_PREFIX):
                candidate = (export_dir / Path(download_url.removeprefix(_HTML_EXPORT_PREFIX)).name).resolve()
                if candidate.parent != export_dir:
                    continue
            if candidate is None or candidate.suffix.lower() not in {".html", ".htm"}:
                continue
            try:
                return candidate.read_text(encoding="utf-8")
            except (OSError, UnicodeError):
                continue
    return None


def _html_revision_context(conv, text: str) -> str:
    if not _is_html_revision_request(text):
        return ""
    html = _latest_html_artifact(conv)
    if not html:
        return ""
    return (
        "Você está editando o último artefato HTML desta conversa. Aplique o pedido "
        "ao documento abaixo e devolva obrigatoriamente o documento HTML completo e "
        "standalone, do <!doctype html> até </html>. Não devolva CSS isolado, trechos, "
        "diff, tutorial ou instruções para o usuário editar manualmente. Preserve o "
        "conteúdo existente que não foi alvo do pedido.\n\n"
        "<artefato_html_anterior>\n"
        f"{html}\n"
        "</artefato_html_anterior>"
    )


def _revision_output_requirement() -> str:
    return (
        "REQUISITO OBRIGATÓRIO DE SAÍDA: responda com o HTML atualizado inteiro, "
        "começando por <!doctype html> e terminando em </html>. Não diga ao usuário "
        "para inserir ou substituir trechos."
    )


def _prepare_session_workspace(conv) -> Path:
    """Materializa uma caixa de chat isolada, rastreável e versionável."""
    workspace = Path(settings.BASE_DIR) / "runtime" / "codex_sessions" / str(conv.id)
    input_dir = workspace / "entrada"
    datasets_dir = input_dir / "datasets"
    documents_dir = input_dir / "documentos"
    datasets_dir.mkdir(parents=True, exist_ok=True)
    documents_dir.mkdir(parents=True, exist_ok=True)
    for dirname in (_WORK_DIRNAME, _OUTPUT_DIRNAME, _EVIDENCE_DIRNAME, _VERSIONS_DIRNAME):
        (workspace / dirname).mkdir(parents=True, exist_ok=True)

    state = conv.state or {}
    evidence_dir = workspace / _EVIDENCE_DIRNAME
    sources = {
        "conversation_id": conv.id,
        "fonte_atual": state.get("athena_last_source") or {},
        "colunas_atuais": state.get("athena_last_columns") or [],
        "datasets_nomeados": sorted((state.get("named_datasets") or {}).keys()),
        "documento_atual": (state.get("documento_atual") or {}).get("filename"),
    }
    (evidence_dir / "fontes.json").write_text(
        json.dumps(sources, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    manifest = {
        "conversation_id": conv.id,
        "estrutura": {
            "entrada": "entrada",
            "trabalho": _WORK_DIRNAME,
            "saida": _OUTPUT_DIRNAME,
            "evidencias": _EVIDENCE_DIRNAME,
            "versoes": _VERSIONS_DIRNAME,
        },
        "fonte_atual": state.get("athena_last_source") or {},
        "colunas_atuais": state.get("athena_last_columns") or [],
        "workbooks": state.get("excel_workbooks") or {},
        "dataset_atual": None,
        "datasets_nomeados": {},
        "documento_atual": None,
        "saidas_geradas": _generated_artifacts_manifest(workspace),
        # Compatibilidade para consumidores anteriores do manifesto.
        "artefatos_gerados": _generated_artifacts_manifest(workspace),
        "evidencias_registradas": _directory_file_manifest(workspace / _EVIDENCE_DIRNAME),
        "versoes_arquivadas": _directory_file_manifest(workspace / _VERSIONS_DIRNAME),
    }
    current = state.get("athena_last_result")
    if current is not None:
        current_path = input_dir / "dataset_atual.json"
        current_path.write_text(json.dumps(current, ensure_ascii=False), encoding="utf-8")
        manifest["dataset_atual"] = str(current_path.relative_to(workspace))

    for index, (name, rows) in enumerate((state.get("named_datasets") or {}).items(), start=1):
        safe_name = re.sub(r"[^A-Za-z0-9_.-]+", "_", str(name)).strip("._")[:80]
        filename = f"{index:03d}_{safe_name or 'dataset'}.json"
        path = datasets_dir / filename
        path.write_text(json.dumps(rows, ensure_ascii=False), encoding="utf-8")
        manifest["datasets_nomeados"][str(name)] = str(path.relative_to(workspace))

    document = state.get("documento_atual") or {}
    document_markdown = document.get("markdown")
    if document_markdown:
        original_name = str(document.get("filename") or "documento")
        safe_name = re.sub(r"[^A-Za-z0-9_.-]+", "_", Path(original_name).stem).strip("._")[:80]
        document_path = documents_dir / f"{safe_name or 'documento'}.md"
        document_path.write_text(str(document_markdown), encoding="utf-8")
        manifest["documento_atual"] = {
            "arquivo": str(document_path.relative_to(workspace)),
            "original": original_name,
            "paginas": document.get("page_count"),
        }

    (workspace / "manifesto_sessao.json").write_text(
        json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    return workspace


def _append_turn_evidence(
    workspace: Path,
    trace_records: dict[str, dict],
    live_plan: list[dict],
    live_plan_explanation: str,
) -> None:
    """Persiste um registro auditável do turno sem alterar arquivos de entrada."""
    evidence_dir = workspace / _EVIDENCE_DIRNAME
    evidence_dir.mkdir(parents=True, exist_ok=True)
    record = {
        "registrado_em_epoch": time.time(),
        "plano": live_plan,
        "explicacao_plano": live_plan_explanation,
        "execucoes": list(trace_records.values()),
    }
    with (evidence_dir / "execucoes.jsonl").open("a", encoding="utf-8") as stream:
        stream.write(json.dumps(record, ensure_ascii=False) + "\n")


def _prompt_with_history(conv, text: str, has_codex_thread: bool) -> str:
    state = conv.state or {}
    workspace = Path(settings.BASE_DIR) / "runtime" / "codex_sessions" / str(conv.id)
    has_generated_artifacts = bool(
        _artifact_snapshot(workspace / _OUTPUT_DIRNAME)
        or _artifact_snapshot(workspace / _LEGACY_ARTIFACTS_DIRNAME)
    )
    has_attached_data = any((
        state.get("athena_last_result") is not None,
        bool(state.get("named_datasets")),
        bool(state.get("excel_workbooks")),
        has_generated_artifacts,
    ))
    hints: list[str] = []
    if has_attached_data:
        hints.append(
            "Há dados anexados disponíveis na sessão. Somente se o pedido atual "
            "exigir esses dados, consulte ../manifesto_sessao.json para localizar o "
            "dataset atual, as abas/datasets nomeados e os artefatos já gerados. "
            "Não mencione essa estrutura "
            "interna na resposta; apresente apenas a análise e suas evidências."
        )

    revision_context = _html_revision_context(conv, text)
    if revision_context:
        hints.append(revision_context)

    session_hint = "\n\n".join(hints)

    def with_hint(value: str) -> str:
        prompted = f"{session_hint}\n\n{value}" if session_hint else value
        if revision_context:
            prompted = f"{prompted}\n\n{_revision_output_requirement()}"
        return prompted

    if has_codex_thread:
        return with_hint(text)
    previous = list(conv.messages.order_by("created_at").values("role", "content"))[-12:]
    if not previous:
        return with_hint(text)
    history = "\n".join(
        f"{('Usuário' if row['role'] == 'user' else 'Assistente')}: {row['content']}"
        for row in previous if row["content"]
    )
    return (
        "Considere este histórico importado da interface antiga apenas como contexto:\n"
        f"<historico>\n{history}\n</historico>\n\n"
        + with_hint(f"Pedido atual do usuário: {text}")
    )


@require_GET
def codex_status(_request):
    from shutil import which

    return JsonResponse({
        "status": "ready" if which("codex") else "unavailable",
        "engine": "codex-app-server",
        "sandbox": "workspace-write-isolado",
        "skills": [
            "auditoria-interna",
            "aws-athena",
            "ciencia-dados",
            "analise-documentos",
            "documentacao-auditoria",
        ],
    })


@require_GET
def codex_execution_detail(_request, execution_id):
    execution = Execution.objects.filter(pk=execution_id).first()
    if execution is None:
        return JsonResponse(
            {"status": "error", "message": "Execução não encontrada."},
            status=404,
        )
    _recover_orphaned_local_executions(execution.conversation_id)
    _expire_unclaimed_worker_execution(execution)
    execution.refresh_from_db()
    return JsonResponse({"status": "success", "execution": _execution_public_payload(execution)})


@csrf_exempt
@require_http_methods(["POST"])
def codex_execution_stop(_request, execution_id):
    execution = Execution.objects.filter(pk=execution_id).first()
    if execution is None:
        return JsonResponse(
            {"status": "error", "message": "Execução não encontrada."},
            status=404,
        )
    _recover_orphaned_local_executions(execution.conversation_id)
    execution.refresh_from_db()
    stopped = request_codex_execution_stop(execution)
    execution.refresh_from_db()
    return JsonResponse({
        "status": "success",
        "stopped": stopped,
        "execution": _execution_public_payload(execution),
    })


_SKILL_SLUG_RE = re.compile(r"^[a-z0-9][a-z0-9-]{0,62}$")
_SKILL_FRONTMATTER_RE = re.compile(r"\A---\s*\n(.*?)\n---\s*\n?", re.DOTALL)


def _skills_root() -> Path:
    return Path(settings.BASE_DIR) / ".agents" / "skills"


def _skill_metadata(slug: str, content: str) -> tuple[str, str]:
    """Lê name/description do frontmatter sem depender de biblioteca YAML."""
    name, description = slug.replace("-", " ").title(), ""
    match = _SKILL_FRONTMATTER_RE.match(content)
    if not match:
        return name, description
    for line in match.group(1).splitlines():
        key, _, value = line.partition(":")
        value = value.strip().strip('"').strip("'")
        if key.strip() == "name" and value:
            name = value
        elif key.strip() == "description":
            description = value
    return name, description


def _skill_payload(prompt_path: Path) -> dict:
    slug = prompt_path.parent.name
    content = prompt_path.read_text(encoding="utf-8")
    name, description = _skill_metadata(slug, content)
    return {"slug": slug, "name": name, "description": description, "prompt": content, "content": content}


@csrf_exempt
@require_http_methods(["GET", "POST"])
def codex_skills(request):
    """Catálogo e criação de skills reais do repositório (.agents/skills)."""
    root = _skills_root()
    if request.method == "GET":
        items = []
        if root.is_dir():
            for prompt_path in sorted(root.glob("*/SKILL.md")):
                try:
                    items.append(_skill_payload(prompt_path))
                except OSError:
                    continue
        return JsonResponse({"skills": items})

    try:
        data = json.loads(request.body or "{}")
        raw_slug = (data.get("slug") or data.get("name") or "").strip().lower()
        slug = re.sub(r"[^a-z0-9]+", "-", unicodedata.normalize("NFKD", raw_slug).encode("ascii", "ignore").decode()).strip("-")
        if not _SKILL_SLUG_RE.fullmatch(slug):
            return JsonResponse({"status": "error", "message": "Use um nome com letras, números e hífens (até 63 caracteres)."}, status=400)
        content = (data.get("prompt") or data.get("content") or "").strip()
        if not content:
            return JsonResponse({"status": "error", "message": "O prompt da skill é obrigatório."}, status=400)
        if len(content) > 100_000:
            return JsonResponse({"status": "error", "message": "O prompt excede o limite de 100 mil caracteres."}, status=400)
        if not _SKILL_FRONTMATTER_RE.match(content):
            name = (data.get("name") or slug).strip()[:120]
            description = (data.get("description") or "").strip()[:500]
            content = f'---\nname: {slug}\ndescription: {json.dumps(description, ensure_ascii=False)}\n---\n\n# {name}\n\n{content}\n'
        prompt_path = root / slug / "SKILL.md"
        prompt_path.parent.mkdir(parents=True, exist_ok=True)
        prompt_path.write_text(content + ("" if content.endswith("\n") else "\n"), encoding="utf-8")
        return JsonResponse({"status": "success", "skill": _skill_payload(prompt_path)})
    except (TypeError, ValueError, OSError) as exc:
        return JsonResponse({"status": "error", "message": str(exc)}, status=400)


@csrf_exempt
@require_http_methods(["DELETE"])
def codex_skill_delete(_request, slug: str):
    if not _SKILL_SLUG_RE.fullmatch(slug):
        return JsonResponse({"status": "error", "message": "Skill inválida."}, status=400)
    skill_dir = _skills_root() / slug
    prompt_path = skill_dir / "SKILL.md"
    if not prompt_path.is_file():
        return JsonResponse({"status": "error", "message": "Skill não encontrada."}, status=404)
    try:
        shutil.rmtree(skill_dir)
        return JsonResponse({"status": "success"})
    except OSError as exc:
        return JsonResponse({"status": "error", "message": str(exc)}, status=400)


@csrf_exempt
def codex_chat_stream(request):
    if request.method != "POST":
        return JsonResponse({"status": "error", "message": "Method not allowed"}, status=405)

    worker_execution = getattr(request, "_codex_worker_execution", None)
    if worker_execution is not None:
        execution_record = Execution.objects.select_related(
            "conversation__agent", "conversation__playbook"
        ).get(pk=worker_execution.pk)
        conv = execution_record.conversation
        data = dict(execution_record.request_payload or {})
        text = (data.get("message") or "").strip()
        if not text:
            return JsonResponse(
                {"status": "error", "message": "Execução sem mensagem persistida."},
                status=400,
            )
        state = dict(conv.state or {})
        old_thread_id = data.get("_old_thread_id") or state.get("codex_thread_id")
        prompt = data.get("_prepared_prompt") or _prompt_with_history(
            conv, text, bool(old_thread_id)
        )
        revision_html = (
            _latest_html_artifact(conv) if _is_html_revision_request(text) else None
        )
    else:
        try:
            data = json.loads(request.body or "{}")
        except json.JSONDecodeError:
            return JsonResponse({"status": "error", "message": "JSON inválido."}, status=400)

        text = (data.get("message") or "").strip()
        if not text:
            return JsonResponse({"status": "error", "message": "Mensagem vazia."}, status=400)

        conv, error = _resolve_conversation(
            data.get("conversation_id"), data.get("agent_slug"), text
        )
        if error:
            return JsonResponse({"status": "error", "message": error}, status=400)

        requested_playbook_id = data.get("playbook_id", _UNSET)
        _apply_playbook(conv, requested_playbook_id)
        conv = Conversation.objects.select_related("agent", "playbook").get(pk=conv.pk)
        if (
            requested_playbook_id is not _UNSET
            and requested_playbook_id is not None
            and str(conv.playbook_id) != str(requested_playbook_id)
        ):
            return JsonResponse({
                "status": "error",
                "code": "playbook_unavailable",
                "message": "Playbook inexistente ou ainda não publicado.",
            }, status=400)
        if conv.playbook_id and conv.playbook.status != "published":
            return JsonResponse({
                "status": "error",
                "code": "playbook_unpublished",
                "message": "Este Playbook está em rascunho. Publique uma versão antes de executar.",
            }, status=409)
        playbook_snapshot = (
            snapshot_for_playbook(conv.playbook) if conv.playbook_id else None
        )
        initial_plan = playbook_plan(playbook_snapshot) if playbook_snapshot else []
        initial_plan_explanation = (
            playbook_plan_explanation(playbook_snapshot) if playbook_snapshot else ""
        )

        _recover_orphaned_local_executions(conv.id)
        active_execution = Execution.objects.filter(
            conversation=conv,
            status__in=Execution.ACTIVE_STATUSES,
        ).first()
        if active_execution is not None:
            return JsonResponse({
                "status": "error",
                "code": "execution_active",
                "message": "Já existe uma execução ativa nesta conversa.",
                "conversation_id": conv.id,
                "execution": _execution_public_payload(active_execution),
            }, status=409)

        state = dict(conv.state or {})
        old_thread_id = state.get("codex_thread_id")
        prompt = _prompt_with_history(conv, text, bool(old_thread_id))
        persisted_payload = {
            **data,
            "conversation_id": conv.id,
            "message": text,
            "_old_thread_id": old_thread_id,
            "_prepared_prompt": prompt,
            "_playbook_snapshot": playbook_snapshot,
        }
        now = timezone.now()
        try:
            with transaction.atomic():
                execution_record = Execution.objects.create(
                    conversation=conv,
                    status="queued",
                    backend="local-worker",
                    request_payload=persisted_payload,
                    last_heartbeat_at=now,
                    plan=initial_plan,
                    plan_explanation=initial_plan_explanation,
                    events=[],
                )
                Message.objects.create(conversation=conv, role="user", content=text)
        except IntegrityError:
            active_execution = Execution.objects.filter(
                conversation=conv,
                status__in=Execution.ACTIVE_STATUSES,
            ).first()
            return JsonResponse({
                "status": "error",
                "code": "execution_active",
                "message": "Já existe uma execução ativa nesta conversa.",
                "conversation_id": conv.id,
                "execution": (
                    _execution_public_payload(active_execution) if active_execution else None
                ),
            }, status=409)

        # O request termina imediatamente. A UI usa o id abaixo para acompanhar
        # os eventos persistidos enquanto o management command executa a tarefa.
        response = StreamingHttpResponse(
            iter((": execution queued\n\n",)),
            content_type="text/event-stream",
        )
        response["Cache-Control"] = "no-cache"
        response["X-Accel-Buffering"] = "no"
        response["X-Conversation-Id"] = str(conv.id)
        response["X-Execution-Id"] = str(execution_record.id)
        return response

    events: "queue.Queue[dict]" = _ExecutionEventQueue(execution_record.id)
    sentinel = object()
    execution = _register_codex_execution(conv.id, str(execution_record.id))
    worker_finished = threading.Event()

    def worker() -> None:
        if execution_record.backend == "local-worker":
            threading.Thread(
                target=_watch_persisted_stop,
                args=(execution_record, execution, worker_finished),
                daemon=True,
            ).start()
        answer_parts: list[str] = []
        approve_all_for_turn = False
        trace_records: dict[str, dict] = {}
        trace_started_at: dict[str, float] = {}
        trace_output: dict[str, str] = {}
        trace_completed: set[str] = set()
        playbook_snapshot = data.get("_playbook_snapshot")
        if not isinstance(playbook_snapshot, dict) or not playbook_snapshot.get("nodes"):
            playbook_snapshot = None
        live_plan: list[dict] = (
            playbook_plan(playbook_snapshot) if playbook_snapshot else []
        )
        live_plan_explanation = (
            playbook_plan_explanation(playbook_snapshot) if playbook_snapshot else ""
        )
        session_workspace: Path | None = None

        def finish_stopped() -> None:
            for item_id, record in trace_records.items():
                if item_id not in trace_completed:
                    record["error"] = record["error"] or "Atividade interrompida pelo usuário."
                    record["result"] = record["result"] or "Execução interrompida."
                    started_at = trace_started_at.get(item_id)
                    if started_at is not None:
                        record["duration_ms"] = round(
                            (time.monotonic() - started_at) * 1000
                        )
            if session_workspace is not None:
                try:
                    _append_turn_evidence(
                        session_workspace,
                        trace_records,
                        live_plan,
                        live_plan_explanation,
                    )
                    _prepare_session_workspace(conv)
                except OSError:
                    pass
            assistant = Message.objects.create(
                conversation=conv,
                role="assistant",
                content="⏹️ Geração interrompida pelo usuário.",
            )
            for record in trace_records.values():
                ToolCall.objects.create(
                    message=assistant,
                    tool_name=record["tool"],
                    args=record["args"],
                    result=record["result"],
                    error=record["error"],
                    duration_ms=record["duration_ms"],
                )
            events.put({
                "type": "done",
                "payload": {
                    "status": "success",
                    "conversation_id": conv.id,
                    "conversation_title": conv.title,
                    "agent_slug": conv.agent.slug if conv.agent else None,
                    "engine": "codex-app-server",
                    "awaiting_human_input": False,
                    "stopped": True,
                    "reply": _message_payload(assistant),
                },
            })

        try:
            Execution.objects.filter(
                pk=execution_record.id,
                status__in=("queued", "starting"),
            ).update(
                status="starting",
                started_at=timezone.now(),
                last_heartbeat_at=timezone.now(),
            )
            if execution.stop_event.is_set():
                raise _CodexExecutionStopped()
            session_workspace = _prepare_session_workspace(conv)
            working_dir = session_workspace / _WORK_DIRNAME
            working_before = _artifact_snapshot(working_dir)
            with CodexAppServer(working_dir) as client:
                _attach_codex_client(execution, client)
                if execution.stop_event.is_set():
                    raise _CodexExecutionStopped()
                events.put({"type": "progress", "stage": "thinking", "icon": "◈", "text": "Conectando à Atena"})
                client.initialize()
                if execution.stop_event.is_set():
                    raise _CodexExecutionStopped()
                thread_id = client.open_thread(old_thread_id)
                Execution.objects.filter(pk=execution_record.id).exclude(
                    status="stopping"
                ).update(
                    status="running",
                    thread_id=thread_id,
                    last_heartbeat_at=timezone.now(),
                )
                state["codex_thread_id"] = thread_id
                state["chat_engine"] = "codex-app-server"
                conv.state = state
                conv.save(update_fields=["state", "updated_at"])
                events.put({"type": "progress", "stage": "thinking", "icon": "🛡️", "text": "Sandbox isolado · trabalho separado das saídas deste chat"})

                def collect_turn(turn_prompt: str) -> str:
                    nonlocal approve_all_for_turn, live_plan, live_plan_explanation

                    if execution.stop_event.is_set():
                        raise _CodexExecutionStopped()

                    def handle_server_request(method: str, params: dict) -> dict:
                        nonlocal approve_all_for_turn
                        is_question = method in {
                            "item/tool/requestUserInput",
                            "tool/requestUserInput",
                        }
                        if approve_all_for_turn and not is_question:
                            events.put({
                                "type": "progress",
                                "stage": "thinking",
                                "icon": "🔓",
                                "text": "Aprovação automática ativa nesta execução",
                            })
                            return _approve_interaction_for_turn(method, params)
                        result = _wait_for_interaction(
                            conv,
                            events,
                            method,
                            params,
                            execution_record=execution_record,
                        )
                        if result.pop("__approve_all_for_turn__", False):
                            approve_all_for_turn = True
                            events.put({
                                "type": "progress",
                                "stage": "thinking",
                                "icon": "🔓",
                                "text": "Todas as próximas ações desta execução foram autorizadas",
                            })
                        return result

                    parts: list[str] = []

                    def remember_turn(turn_id: str) -> None:
                        Execution.objects.filter(pk=execution_record.id).update(
                            turn_id=turn_id,
                            last_heartbeat_at=timezone.now(),
                        )

                    for event in client.turn(
                        thread_id,
                        turn_prompt,
                        server_request_handler=handle_server_request,
                        turn_started_handler=remember_turn,
                    ):
                        if event["type"] == "delta" and event.get("phase") != "commentary":
                            parts.append(event["text"])
                        elif event["type"] == "plan" and not playbook_snapshot:
                            live_plan = _normalize_live_plan(event.get("plan") or [])
                            live_plan_explanation = _sanitize_trace_text(
                                event.get("explanation"), 1200
                            )
                            events.put({
                                "type": "plan",
                                "explanation": live_plan_explanation,
                                "plan": live_plan,
                            })
                        elif event["type"] == "activity":
                            item = event.get("item") or {}
                            item_id = str(item.get("id") or "")
                            phase = event.get("phase")
                            if phase == "started":
                                record = _codex_trace_record(item)
                                if record:
                                    trace_records[item_id] = record
                                    trace_started_at[item_id] = time.monotonic()
                                    events.put({
                                        "type": "progress",
                                        "stage": "tool",
                                        "agent": "codex",
                                        "tool": record["tool"],
                                        "tool_call_id": item_id,
                                        "args": record["args"],
                                        "icon": record["icon"],
                                        "text": record["label"],
                                    })
                            elif phase == "completed":
                                started_at = trace_started_at.get(item_id)
                                elapsed_ms = (
                                    round((time.monotonic() - started_at) * 1000)
                                    if started_at is not None else 0
                                )
                                record = _codex_trace_record(
                                    item,
                                    trace_output.get(item_id, ""),
                                    elapsed_ms,
                                )
                                if record:
                                    # Alguns clientes podem receber somente o completed;
                                    # ainda assim criamos o nó antes de encerrá-lo.
                                    if item_id not in trace_records:
                                        events.put({
                                            "type": "progress",
                                            "stage": "tool",
                                            "agent": "codex",
                                            "tool": record["tool"],
                                            "tool_call_id": item_id,
                                            "args": record["args"],
                                            "icon": record["icon"],
                                            "text": record["label"],
                                        })
                                    trace_records[item_id] = record
                                    trace_completed.add(item_id)
                                    events.put({
                                        "type": "progress",
                                        "stage": "tool_result",
                                        "agent": "codex",
                                        "tool": record["tool"],
                                        "tool_call_id": item_id,
                                        "error": record["error"],
                                        "duration_ms": record["duration_ms"],
                                        "result_preview": _sanitize_trace_text(
                                            record["result"], _TRACE_RESULT_PREVIEW_LIMIT
                                        ),
                                    })
                                    advanced_plan = (
                                        None if playbook_snapshot else _advance_live_plan(live_plan)
                                    )
                                    if advanced_plan is not None:
                                        live_plan = advanced_plan
                                        events.put({
                                            "type": "plan",
                                            "explanation": live_plan_explanation,
                                            "plan": live_plan,
                                        })
                        elif event["type"] == "activity_output":
                            item_id = str(event.get("item_id") or "")
                            if item_id:
                                current = trace_output.get(item_id, "")
                                trace_output[item_id] = _sanitize_trace_text(
                                    current + str(event.get("text") or "")
                                )
                        elif event["type"] == "completed" and event.get("status") != "completed":
                            if execution.stop_event.is_set() or event.get("status") == "interrupted":
                                raise _CodexExecutionStopped()
                            raise CodexAppServerError(
                                str(event.get("error") or "Turno não concluído.")
                            )
                    if execution.stop_event.is_set():
                        raise _CodexExecutionStopped()
                    return "".join(parts).strip()

                if playbook_snapshot:
                    policy = normalize_execution_policy(
                        playbook_snapshot.get("execution_policy")
                    )
                    stages = ordered_playbook_nodes(playbook_snapshot)
                    if not stages:
                        stages = [
                            node
                            for node in (playbook_snapshot.get("nodes") or [])
                            if isinstance(node, dict) and node.get("is_root")
                        ][:1]
                    stage_results: list[tuple[str, str]] = []
                    total_stages = len(stages)

                    for stage_index, stage in enumerate(stages):
                        stage_name = str(
                            stage.get("name") or stage.get("slug") or "Etapa"
                        )
                        live_plan = set_plan_stage(
                            live_plan, stage_index, "inProgress"
                        )
                        events.put({
                            "type": "plan",
                            "explanation": live_plan_explanation,
                            "plan": live_plan,
                        })
                        events.put({
                            "type": "progress",
                            "stage": "playbook_stage",
                            "icon": str(stage.get("icon") or "◆"),
                            "text": f"Etapa {stage_index + 1}/{total_stages} · {stage_name}",
                        })

                        requires_gate = bool(stage.get("requires_approval")) or bool(
                            policy["require_stage_confirmation"]
                        )
                        if requires_gate:
                            gate = _wait_for_interaction(
                                conv,
                                events,
                                "item/tool/requestUserInput",
                                {
                                    "questions": [{
                                        "id": "stage_gate",
                                        "header": "Próxima etapa",
                                        "question": (
                                            f"Autoriza executar a etapa "
                                            f"{stage_index + 1}/{total_stages}: {stage_name}?"
                                        ),
                                        "options": [
                                            {
                                                "label": "Aprovar",
                                                "description": "Executar esta etapa agora.",
                                            },
                                            {
                                                "label": "Cancelar",
                                                "description": "Interromper o Playbook com segurança.",
                                            },
                                        ],
                                    }],
                                },
                                execution_record=execution_record,
                            )
                            gate_answers = (
                                gate.get("answers", {})
                                .get("stage_gate", {})
                                .get("answers", [])
                            )
                            gate_choice = str(gate_answers[0]).lower() if gate_answers else ""
                            if "aprovar" not in gate_choice:
                                raise _CodexExecutionStopped()

                        handoff_context = prompt
                        if stage_results:
                            prior = "\n\n".join(
                                f"Handoff de {name}:\n{result[:12000]}"
                                for name, result in stage_results
                            )
                            handoff_context = f"{prompt}\n\n{prior}"
                        stage_prompt = playbook_stage_prompt(
                            playbook_snapshot,
                            stage,
                            stage_index,
                            total_stages,
                            handoff_context,
                        )
                        max_attempts = int(stage.get("max_retries") or 0) + 1
                        stage_answer = ""
                        stage_error: Exception | None = None
                        for attempt in range(max_attempts):
                            try:
                                stage_answer = collect_turn(stage_prompt)
                                stage_error = None
                                break
                            except CodexAppServerError as exc:
                                stage_error = exc
                                if attempt + 1 < max_attempts:
                                    events.put({
                                        "type": "progress",
                                        "stage": "playbook_retry",
                                        "icon": "↻",
                                        "text": (
                                            f"Repetindo {stage_name} "
                                            f"({attempt + 2}/{max_attempts})"
                                        ),
                                    })
                        if stage_error is not None:
                            live_plan = set_plan_stage(live_plan, stage_index, "failed")
                            events.put({
                                "type": "plan",
                                "explanation": live_plan_explanation,
                                "plan": live_plan,
                            })
                            can_continue = (
                                str(stage.get("on_error") or "stop") == "continue"
                                or not policy["stop_on_error"]
                            )
                            if not can_continue:
                                raise stage_error
                            stage_results.append((
                                stage_name,
                                f"Etapa falhou e o fluxo prosseguiu: {stage_error}",
                            ))
                            continue

                        live_plan = set_plan_stage(live_plan, stage_index, "completed")
                        events.put({
                            "type": "plan",
                            "explanation": live_plan_explanation,
                            "plan": live_plan,
                        })
                        stage_results.append((stage_name, stage_answer))

                    if policy["final_synthesis"] and len(stage_results) > 1:
                        events.put({
                            "type": "progress",
                            "stage": "playbook_synthesis",
                            "icon": "◇",
                            "text": "Consolidando os resultados do Playbook",
                        })
                        answer = collect_turn(
                            playbook_synthesis_prompt(
                                playbook_snapshot,
                                text,
                                stage_results,
                            )
                        )
                    else:
                        answer = stage_results[-1][1] if stage_results else ""
                else:
                    answer = collect_turn(prompt)
                if execution.stop_event.is_set():
                    raise _CodexExecutionStopped()
                if revision_html and _split_html_response(answer)[1] is None:
                    events.put({
                        "type": "progress",
                        "stage": "thinking",
                        "icon": "◇",
                        "text": "Consolidando a nova versão completa do HTML",
                    })
                    repair_prompt = (
                        "A resposta anterior ficou incompleta porque trouxe apenas "
                        "instruções ou fragmentos. Refaça agora o pedido original: "
                        f"{text}\n\n{_revision_output_requirement()}\n\n"
                        "Use este documento como base e preserve o restante:\n"
                        "<artefato_html_anterior>\n"
                        f"{revision_html}\n"
                        "</artefato_html_anterior>"
                    )
                    answer = collect_turn(repair_prompt)

                answer_parts.append(answer)

            if execution.stop_event.is_set():
                raise _CodexExecutionStopped()
            for item_id, record in trace_records.items():
                if item_id not in trace_completed:
                    started_at = trace_started_at.get(item_id)
                    if started_at is not None:
                        record["duration_ms"] = round(
                            (time.monotonic() - started_at) * 1000
                        )
                    record["result"] = record["result"] or "Atividade concluída com o turno."

            answer = "".join(answer_parts).strip()
            if not answer:
                answer = "A Atena concluiu o turno sem produzir uma mensagem de texto."
            generated_attachments = _collect_generated_artifacts(
                working_dir,
                working_before,
                conv.id,
            )
            if generated_attachments:
                answer = _strip_codex_file_citations(answer)
            answer, html_attachments = _materialize_html_response(answer, conv.id)
            attachments = generated_attachments + html_attachments
            _append_turn_evidence(
                session_workspace,
                trace_records,
                live_plan,
                live_plan_explanation,
            )
            # Atualiza o índice somente após saídas e evidências existirem.
            _prepare_session_workspace(conv)
            assistant = Message.objects.create(
                conversation=conv,
                role="assistant",
                content=answer,
                attachments=attachments,
            )
            for record in trace_records.values():
                ToolCall.objects.create(
                    message=assistant,
                    tool_name=record["tool"],
                    args=record["args"],
                    result=record["result"],
                    error=record["error"],
                    duration_ms=record["duration_ms"],
                )
            events.put({
                "type": "done",
                "payload": {
                    "status": "success",
                    "conversation_id": conv.id,
                    "conversation_title": conv.title,
                    "agent_slug": conv.agent.slug if conv.agent else None,
                    "engine": "codex-app-server",
                    "awaiting_human_input": False,
                    "reply": _message_payload(assistant),
                },
            })
        except _CodexExecutionStopped:
            finish_stopped()
        except Exception as exc:
            if execution.stop_event.is_set():
                finish_stopped()
            else:
                events.put({"type": "error", "message": f"Atena: {exc}"})
        finally:
            from django.db import connection

            worker_finished.set()
            _unregister_codex_execution(execution)
            Conversation.objects.filter(pk=conv.id).update(awaiting_human_input=False)
            connection.close()
            events.put(sentinel)

    def stream():
        yield ": stream start\n\n"
        threading.Thread(target=worker, daemon=True).start()
        while True:
            try:
                event = events.get(timeout=15)
            except queue.Empty:
                yield ": keep-alive\n\n"
                continue
            if event is sentinel:
                break
            yield f"data: {json.dumps(event, ensure_ascii=False)}\n\n"

    response = StreamingHttpResponse(stream(), content_type="text/event-stream")
    response["Cache-Control"] = "no-cache"
    response["X-Accel-Buffering"] = "no"
    response["X-Conversation-Id"] = str(conv.id)
    response["X-Execution-Id"] = str(execution_record.id)
    return response


def run_queued_codex_execution(execution: Execution | str) -> Execution:
    """Executa uma tarefa persistida consumindo o mesmo pipeline do SSE.

    O management command chama esta função em um processo separado. Consumir
    o stream internamente mantém compatibilidade com o contrato de eventos sem
    manter o request original do navegador aberto.
    """
    if not isinstance(execution, Execution):
        execution = Execution.objects.get(pk=execution)
    if execution.backend != "local-worker":
        raise ValueError("A execução não pertence ao worker local.")
    if execution.status in Execution.TERMINAL_STATUSES:
        return execution

    request = HttpRequest()
    request.method = "POST"
    request._body = json.dumps(execution.request_payload or {}).encode("utf-8")
    request.META["CONTENT_TYPE"] = "application/json"
    request._codex_worker_execution = execution
    response = codex_chat_stream(request)
    if not getattr(response, "streaming", False):
        try:
            payload = json.loads(response.content or "{}")
            message = payload.get("message") or "Falha ao iniciar execução persistida."
        except (AttributeError, json.JSONDecodeError):
            message = "Falha ao iniciar execução persistida."
        raise RuntimeError(message)
    try:
        for _chunk in response.streaming_content:
            pass
    finally:
        response.close()
    execution.refresh_from_db()
    return execution
